from datetime import datetime
import openpyxl
import os

def createFolder(directory):
    try:
        if not os.path.exists(directory):
            os.makedirs(directory)
    except OSError:
        print('Error: Creating directory. ' + directory)

itemname = [] #항목
company = [] #협력사명
item = [] #값
item.append([]) #
value = []

dir = os.getcwd()+'/'+datetime.today().strftime("%m%d") # ex) dir = 'C:\Users\사용자 폴더\Desktop'
file_name = '0215 전체발주건' # ex) file_name = '1224 전체발주'

wb_load = openpyxl.load_workbook(file_name+'.xlsx')

sheet_load = wb_load.active

#항목 저장
for i in range(1,17):
    itemname.append(sheet_load.cell(row = 1 , column = i).value)
print(itemname)


row = 2
while (sheet_load.cell(row, column=3).value):
    name = sheet_load.cell(row, column=16).value
    if (name in company):
        for i in range(1,17):
            value.append(sheet_load.cell(row = row , column = i).value)

        item[company.index(name)].append(value)
        row = row+1
        value = []
    else:
        company.append(name)
        item.append([])
        item[company.index(name)].append(itemname)

createFolder(dir)

for i in range(len(company)):
    wb_save = openpyxl.Workbook()
    sheet_save = wb_save.active
    for j in range(len(item[i])):
        li = item[i][j]
        sheet_save.append(li)
    wb_save.save(dir+'/'+datetime.today().strftime("%m%d")+" "+company[i] + ".xlsx")